"""
Логика выгрузки данных о пользователях в Excel файл.

Создаёт xlsx файл с информацией о всех пользователях системы.
"""

from datetime import datetime
from io import BytesIO
from typing import Optional

from aiogram.types import BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.database import User
from app.database.utils import now_msk
from app.services.const import USER_STATUS_NAMES


# Константы для конфигурации
HEADERS = [
    "ID",
    "Username",
    "Статус",
    "Email",
    "Имя",
    "Возраст",
    "Интересы",
    "Дата регистрации",
    "Последний мэтч",
]

COLUMN_WIDTHS = {
    "A": 8,  # ID
    "B": 15,  # Username
    "C": 12,  # Статус
    "D": 25,  # Email
    "E": 20,  # Имя
    "F": 10,  # Возраст
    "H": 30,  # Интересы
    "I": 20,  # Дата регистрации
    "J": 20,  # Последний мэтч
}

TEXT_COLUMNS = [2, 3, 4, 5, 6, 7, 8, 9]  # Номера текстовых столбцов


async def _get_users(session: AsyncSession) -> list[User]:
    """
    Получает всех пользователей из базы данных, отсортированных по дате регистрации.

    Args:
        session (AsyncSession): сессия БД.

    Returns:
        list[User]: список пользователей.
    """
    result = await session.execute(select(User).order_by(User.registered_at))
    return result.scalars().all()


def _create_workbook() -> tuple[Workbook, Worksheet]:
    """
    Создаёт новую рабочую книгу Excel и лист.

    Returns:
        tuple[Workbook, Worksheet]: кортеж из рабочей книги и активного листа.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    return wb, ws


def _write_headers(ws: Worksheet) -> None:
    """
    Записывает заголовки столбцов в лист Excel с форматированием.

    Args:
        ws (Worksheet): рабочий лист Excel.
    """
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def _format_datetime(dt: Optional[datetime]) -> str:
    """
    Форматирует дату и время в строку.

    Args:
        dt (Optional[datetime]): дата для форматирования.

    Returns:
        str: отформатированная дата или пустая строка.
    """
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    return ""


def _extract_interests(user: User) -> str:
    """
    Извлекает интересы пользователя из JSON поля.

    Args:
        user (User): объект пользователя.

    Returns:
        str: строка с интересами, разделёнными запятыми.
    """
    if user.interests_json and isinstance(user.interests_json, dict):
        interests_list = user.interests_json.get("interests", [])
        if isinstance(interests_list, list):
            return ", ".join(str(i) for i in interests_list)
    return ""


def _get_user_row_data(user: User) -> list:
    """
    Формирует список данных пользователя для записи в строку Excel.

    Args:
        user (User): объект пользователя.

    Returns:
        list: список значений для строки Excel.
    """
    interests = _extract_interests(user)
    status_name = USER_STATUS_NAMES.get(user.status, user.status)

    return [
        user.id,
        user.username,
        status_name,
        user.email,
        user.name or "",
        user.age or "",
        interests,
        _format_datetime(user.registered_at),
        _format_datetime(user.last_match_at),
    ]


def _write_user_data(ws: Worksheet, users: list[User]) -> None:
    """
    Записывает данные пользователей в лист Excel.

    Args:
        ws (Worksheet): рабочий лист Excel.
        users (list[User]): список пользователей.
    """
    for row_num, user in enumerate(users, start=2):
        row_data = _get_user_row_data(user)

        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)

            # Выравнивание для текстовых полей
            if col_num in TEXT_COLUMNS:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:  # Числовые столбцы (ID, Возраст)
                cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_formatting(ws: Worksheet) -> None:
    """
    Применяет форматирование к листу Excel (ширина столбцов, фиксация заголовков).

    Args:
        ws (Worksheet): рабочий лист Excel.
    """
    # Настраиваем ширину столбцов
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Фиксируем первую строку (заголовки)
    ws.freeze_panes = "A2"


def _save_to_bytes(wb: Workbook) -> BytesIO:
    """
    Сохраняет рабочую книгу Excel в байтовый поток.

    Args:
        wb (Workbook): рабочая книга Excel.

    Returns:
        BytesIO: байтовый поток с содержимым xlsx файла.
    """
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def export_users_to_excel(
    session_factory: async_sessionmaker[AsyncSession],
) -> BytesIO:
    """
    Экспортирует данные о всех пользователях в Excel файл.

    Создаёт xlsx файл с информацией о пользователях:
    - ID, Username
    - Статус
    - Email
    - Данные анкеты (имя, возраст, интересы)
    - Даты регистрации и последнего мэтча

    Args:
        session_factory (async_sessionmaker[AsyncSession]): фабрика сессий БД.

    Returns:
        BytesIO: байтовый поток с содержимым xlsx файла.
    """
    async with session_factory() as session:
        users = await _get_users(session)

        wb, ws = _create_workbook()
        # Запись заголовков и данных пользователей
        _write_headers(ws)
        _write_user_data(ws, users)
        # Применение форматирования
        _apply_formatting(ws)

        # Бинарник Excel
        excel_bytes = _save_to_bytes(wb)
        excel_bytes.seek(0)  # Перенос потока в начало файла

        # Получение сегодняшней даты в формате YYYYMMDD
        today = now_msk().strftime("%Y%m%d")

        document = BufferedInputFile(excel_bytes.read(), filename=f"users-{today}.xlsx")

        return document
