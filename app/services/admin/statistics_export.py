"""
Логика экспорта статистики в Excel файл.

Создаёт xlsx файл с двумя листами:
- Общая статистика (метрики)
- Статистика по периодам (метрики)
"""

from io import BytesIO

from aiogram.types import BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.database.utils import now_msk
from app.services.admin.statistics import (
    get_general_statistics,
    get_period_statistics,
    PERIOD_LABELS,
)


# Стили
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ----------------------------- Лист "Статистика" ----------------------------- #


STATS_HEADERS = ["Период", "Всего мэтчей", "Успешных мэтчей", "Ср. Jaccard-коэффициент"]
STATS_COLUMN_WIDTHS = {"A": 20, "B": 18, "C": 20, "D": 25}


def _write_stats_headers(ws: Worksheet) -> None:
    """Записывает заголовки листа статистики."""
    for col_num, header in enumerate(STATS_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT


async def _write_stats_data(ws: Worksheet, session: AsyncSession) -> None:
    """Записывает данные статистики по периодам."""
    periods = ["7_days", "30_days", "6_months", "all_time"]

    for row_num, period in enumerate(periods, start=2):
        period_stats = await get_period_statistics(session, period)

        period_label = PERIOD_LABELS.get(period, period)
        total = period_stats.get("total_matches", 0)
        successful = period_stats.get("successful_matches", 0)
        avg_jaccard = period_stats.get("average_jaccard_score")
        jaccard_str = f"{avg_jaccard:.3f}" if avg_jaccard is not None else "—"

        row_data = [period_label, total, successful, jaccard_str]
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_stats_formatting(ws: Worksheet) -> None:
    """Применяет форматирование к листу статистики."""
    for col_letter, width in STATS_COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"


# ----------------------------- Лист "Общая статистика" ----------------------------- #


GENERAL_STATS_HEADERS = ["Метрика", "Значение"]
GENERAL_STATS_COLUMN_WIDTHS = {"A": 35, "B": 15}


def _write_general_stats_headers(ws: Worksheet) -> None:
    """Записывает заголовки листа общей статистики."""
    for col_num, header in enumerate(GENERAL_STATS_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT


async def _write_general_stats_data(ws: Worksheet, session: AsyncSession) -> None:
    """Записывает данные общей статистики."""
    general = await get_general_statistics(session)

    metrics = [
        ("Активных пользователей", general.get("active_users_count", 0)),
        ("Заблокированных пользователей", general.get("blocked_users_count", 0)),
    ]

    for row_num, (metric_name, value) in enumerate(metrics, start=2):
        ws.cell(row=row_num, column=1, value=metric_name).alignment = Alignment(
            vertical="center"
        )
        ws.cell(row=row_num, column=2, value=value).alignment = Alignment(
            horizontal="center", vertical="center"
        )


def _apply_general_stats_formatting(ws: Worksheet) -> None:
    """Применяет форматирование к листу общей статистики."""
    for col_letter, width in GENERAL_STATS_COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"


# ----------------------------- Основная функция экспорта ----------------------------- #


def _save_to_bytes(wb: Workbook) -> BytesIO:
    """Сохраняет рабочую книгу Excel в байтовый поток."""
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def export_statistics_to_excel(
    session_factory: async_sessionmaker[AsyncSession],
) -> BufferedInputFile:
    """
    Экспортирует статистику в Excel файл.

    Создаёт xlsx файл с двумя листами:
    - Общая статистика (активные/заблокированные пользователи)
    - Статистика по периодам (мэтчи по времени)

    Args:
        session_factory (async_sessionmaker[AsyncSession]): фабрика сессий БД.

    Returns:
        BufferedInputFile: файл для отправки в Telegram.
    """
    async with session_factory() as session:
        wb = Workbook()

        # Лист 1: Общая статистика
        ws_general = wb.active
        ws_general.title = "Общая статистика"
        _write_general_stats_headers(ws_general)
        await _write_general_stats_data(ws_general, session)
        _apply_general_stats_formatting(ws_general)

        # Лист 2: Статистика по периодам
        ws_stats = wb.create_sheet("Статистика по периодам")
        _write_stats_headers(ws_stats)
        await _write_stats_data(ws_stats, session)
        _apply_stats_formatting(ws_stats)

        # Сохранение в байты
        excel_bytes = _save_to_bytes(wb)
        excel_bytes.seek(0)

        # Формирование имени файла
        today = now_msk().strftime("%Y%m%d")
        document = BufferedInputFile(
            excel_bytes.read(), filename=f"statistics-{today}.xlsx"
        )

        return document
